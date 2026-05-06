"""
Excel Export module — Streamlit page.
Exports replenishment signals, buffer parameters, and demand/supply data to .xlsx.
"""

import io
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from database.db import (
    get_session, Item, Buffer, DemandEntry, SupplyEntry,
    BufferAdjustment, BomLine,
)
from database.auth import get_company_id


STATUS_FILL = {
    "red":     PatternFill("solid", fgColor="FADBD8"),
    "yellow":  PatternFill("solid", fgColor="FDEBD0"),
    "green":   PatternFill("solid", fgColor="D5F5E3"),
    "unknown": PatternFill("solid", fgColor="F2F3F4"),
}

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def show():
    st.header("Export to Excel")
    st.caption("Generate Excel reports for replenishment signals and buffer data.")

    tab_signals, tab_buffers, tab_demand, tab_supply, tab_adj, tab_bom, tab_mv = st.tabs([
        "Replenishment Signals", "Buffer Parameters", "Demand Entries", "Supply Entries",
        "Buffer Adjustments", "BOM Lines", "Model Velocity",
    ])

    with tab_signals:
        _export_signals()

    with tab_buffers:
        _export_buffer_params()

    with tab_demand:
        _export_demand()

    with tab_supply:
        _export_supply()

    with tab_adj:
        _export_adjustments()

    with tab_bom:
        _export_bom()

    with tab_mv:
        _export_model_velocity()


# ---------------------------------------------------------------------------
# Replenishment Signals export
# ---------------------------------------------------------------------------

def _export_signals():
    st.subheader("Replenishment Signals Export")
    st.caption("Exports all items with their current buffer status and suggested order quantities.")

    if st.button("Generate Signals Report", type="primary"):
        wb = _build_signals_workbook()
        buf = _wb_to_bytes(wb)
        filename = f"DDMRP_Signals_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label="Download Signals Report",
            data=buf,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def _build_signals_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Replenishment Signals"

    headers = [
        "Status", "Exec Band", "Status %",
        "Part Number", "Description", "Category",
        "On Hand", "On Order (open)", "Net Flow Position",
        "Top of Red", "Top of Yellow", "Top of Green",
        "Suggested Order Qty", "ADU", "DLT (days)", "Last Calculated",
    ]
    _write_header_row(ws, headers, row=1)

    session = get_session()
    try:
        items = session.query(Item).filter(Item.company_id == get_company_id()).order_by(Item.part_number).all()
        buffers = {b.item_id: b for b in session.query(Buffer).all()}

        priority = {"red": 0, "yellow": 1, "green": 2}
        items_sorted = sorted(
            items,
            key=lambda it: priority.get(buffers[it.id].status if it.id in buffers else "green", 2)
        )

        for row_idx, item in enumerate(items_sorted, start=2):
            buf = buffers.get(item.id)
            status = buf.status if buf else "unknown"
            fill = STATUS_FILL.get(status, STATUS_FILL["unknown"])

            exec_band = (buf.execution_color or "green") if buf else "green"
            status_pct = round((buf.buffer_status_pct or 0.0) * 100, 0) if buf else 0
            row_data = [
                status.upper(),
                exec_band,
                f"{status_pct:.0f}%",
                item.part_number,
                item.description,
                item.category,
                item.on_hand,
                buf.net_flow_position - item.on_hand if buf else 0,
                round(buf.net_flow_position, 2) if buf else 0,
                round(buf.top_of_red, 2) if buf else 0,
                round(buf.top_of_yellow, 2) if buf else 0,
                round(buf.top_of_green, 2) if buf else 0,
                round(buf.suggested_order_qty, 2) if buf else 0,
                item.adu,
                item.dlt,
                buf.last_calculated.strftime("%Y-%m-%d %H:%M") if buf and buf.last_calculated else "",
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
    finally:
        session.close()

    _autofit_columns(ws)
    return wb


# ---------------------------------------------------------------------------
# Buffer Parameters export
# ---------------------------------------------------------------------------

def _export_buffer_params():
    st.subheader("Buffer Parameters Export")
    st.caption("Exports all DDMRP parameters and calculated zone sizes for every item.")

    if st.button("Generate Buffer Parameters Report", type="primary"):
        wb = _build_params_workbook()
        buf = _wb_to_bytes(wb)
        filename = f"DDMRP_BufferParams_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label="Download Buffer Parameters",
            data=buf,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def _build_params_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Buffer Parameters"

    headers = [
        "Part Number", "Description", "Category", "UoM",
        "Item Type", "Buffer Profile",
        "ADU", "DLT (days)", "Lead Time Factor", "Variability Factor",
        "Min Order Qty", "Order Cycle (days)",
        "Spike Horizon (days)", "Spike Threshold Factor",
        "Unit Cost (€)", "Ordering Cost (€)", "Holding Cost %",
        "Red Zone Base", "Red Zone Safety", "Red Zone (TOR)",
        "Yellow Zone", "Green Zone",
        "Top of Red", "Top of Yellow", "Top of Green",
        "Avg Inventory Target", "Avg Order Freq (days)",
        "Safety Days", "Avg Active Orders",
        "Computed DLT (BOM)", "Exec Band", "Buffer Status %",
    ]
    _write_header_row(ws, headers, row=1)

    from modules.buffer_engine import calculate_zones
    from modules.bom_engine import compute_all_dlt
    dlt_map = {r.item_id: r for r in compute_all_dlt(company_id=get_company_id())}
    session = get_session()
    try:
        items   = session.query(Item).filter(Item.company_id == get_company_id()).order_by(Item.part_number).all()
        buffers = {b.item_id: b for b in session.query(Buffer).all()}
        for row_idx, item in enumerate(items, start=2):
            z = calculate_zones(item)
            profile_name = item.buffer_profile.name if item.buffer_profile else ""
            dlt_r = dlt_map.get(item.id)
            buf   = buffers.get(item.id)
            row_data = [
                item.part_number, item.description, item.category, item.unit_of_measure,
                item.item_type or "P", profile_name,
                item.adu, item.dlt, item.lead_time_factor, item.variability_factor,
                item.min_order_qty, item.order_cycle,
                item.spike_horizon_days if item.spike_horizon_days else "",
                item.spike_threshold_factor if item.spike_threshold_factor else "",
                round(item.unit_cost or 0.0, 2),
                round(item.ordering_cost or 0.0, 2),
                round((item.holding_cost_pct or 0.0) * 100, 2),
                round(z.red_zone_base, 2), round(z.red_zone_safety, 2), round(z.red_zone, 2),
                round(z.yellow_zone, 2), round(z.green_zone, 2),
                round(z.top_of_red, 2), round(z.top_of_yellow, 2), round(z.top_of_green, 2),
                round(z.avg_inventory_target, 2),
                round(z.avg_order_frequency_days, 2),
                round(z.safety_days, 2),
                round(z.avg_active_orders, 2),
                round(dlt_r.computed_dlt, 2) if dlt_r else item.dlt,
                (buf.execution_color or "green") if buf else "green",
                f"{round((buf.buffer_status_pct or 0.0) * 100, 0):.0f}%" if buf else "0%",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
    finally:
        session.close()

    _autofit_columns(ws)
    return wb


# ---------------------------------------------------------------------------
# Demand export
# ---------------------------------------------------------------------------

def _export_demand():
    st.subheader("Demand Entries Export")
    if st.button("Generate Demand Report", type="primary"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Demand Entries"
        headers = ["ID", "Part Number", "Description", "Type",
                   "Quantity", "Demand Date", "Reference", "Notes"]
        _write_header_row(ws, headers)

        session = get_session()
        try:
            entries = (
                session.query(DemandEntry, Item)
                .join(Item)
                .order_by(DemandEntry.demand_date)
                .all()
            )
            for row_idx, (e, it) in enumerate(entries, start=2):
                row_data = [
                    e.id, it.part_number, it.description, e.demand_type,
                    e.quantity, e.demand_date.strftime("%Y-%m-%d"),
                    e.order_reference, e.notes,
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=val).border = THIN_BORDER
        finally:
            session.close()

        _autofit_columns(ws)
        buf = _wb_to_bytes(wb)
        st.download_button(
            "Download Demand Report",
            data=buf,
            file_name=f"DDMRP_Demand_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ---------------------------------------------------------------------------
# Supply export
# ---------------------------------------------------------------------------

def _export_supply():
    st.subheader("Supply Entries Export")
    if st.button("Generate Supply Report", type="primary"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Supply Entries"
        headers = ["ID", "Part Number", "Description", "Type",
                   "Quantity", "Due Date", "Reference", "Notes"]
        _write_header_row(ws, headers)

        session = get_session()
        try:
            entries = (
                session.query(SupplyEntry, Item)
                .join(Item)
                .order_by(SupplyEntry.due_date)
                .all()
            )
            for row_idx, (e, it) in enumerate(entries, start=2):
                row_data = [
                    e.id, it.part_number, it.description,
                    e.supply_type.replace("_", " ").title(),
                    e.quantity, e.due_date.strftime("%Y-%m-%d"),
                    e.order_reference, e.notes,
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=val).border = THIN_BORDER
        finally:
            session.close()

        _autofit_columns(ws)
        buf = _wb_to_bytes(wb)
        st.download_button(
            "Download Supply Report",
            data=buf,
            file_name=f"DDMRP_Supply_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ---------------------------------------------------------------------------
# Buffer Adjustments export
# ---------------------------------------------------------------------------

def _export_adjustments():
    st.subheader("Buffer Adjustments Export")
    st.caption("Exports all DAF / LTAF / ZAF planned adjustments.")

    if st.button("Generate Adjustments Report", type="primary", key="exp_adj"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Buffer Adjustments"
        headers = [
            "ID", "Part Number", "Description",
            "Start Date", "End Date",
            "DAF", "LTAF", "Red ZAF", "Yellow ZAF", "Green ZAF", "Note",
        ]
        _write_header_row(ws, headers)

        session = get_session()
        try:
            rows = (
                session.query(BufferAdjustment, Item)
                .join(Item, BufferAdjustment.item_id == Item.id)
                .order_by(BufferAdjustment.start_date)
                .all()
            )
            for i, (adj, it) in enumerate(rows, start=2):
                row_data = [
                    adj.id, it.part_number, it.description,
                    adj.start_date.strftime("%Y-%m-%d") if adj.start_date else "",
                    adj.end_date.strftime("%Y-%m-%d") if adj.end_date else "(open)",
                    adj.daf, adj.ltaf,
                    adj.red_zaf, adj.yellow_zaf, adj.green_zaf,
                    adj.note or "",
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=i, column=col_idx, value=val).border = THIN_BORDER
        finally:
            session.close()

        _autofit_columns(ws)
        buf = _wb_to_bytes(wb)
        st.download_button(
            "Download Adjustments Report", data=buf,
            file_name=f"DDMRP_Adjustments_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ---------------------------------------------------------------------------
# BOM Lines export
# ---------------------------------------------------------------------------

def _export_bom():
    st.subheader("BOM Lines Export")
    st.caption("Exports all Bill of Materials lines with computed DLT.")

    if st.button("Generate BOM Report", type="primary", key="exp_bom"):
        from modules.bom_engine import compute_all_dlt
        dlt_map = {r.item_id: r for r in compute_all_dlt(company_id=get_company_id())}

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM Lines"
        headers = [
            "ID", "Parent Part", "Parent Description",
            "Child Part", "Child Description", "Qty", "Note",
        ]
        _write_header_row(ws, headers)

        session = get_session()
        try:
            items = {it.id: it for it in session.query(Item).filter(Item.company_id == get_company_id()).all()}
            lines = session.query(BomLine).order_by(
                BomLine.parent_item_id, BomLine.child_item_id).all()
            for i, l in enumerate(lines, start=2):
                p = items.get(l.parent_item_id)
                c = items.get(l.child_item_id)
                row_data = [
                    l.id,
                    p.part_number if p else "?", p.description if p else "",
                    c.part_number if c else "?", c.description if c else "",
                    l.qty, l.note or "",
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=i, column=col_idx, value=val).border = THIN_BORDER
        finally:
            session.close()

        # Second sheet: computed DLT summary
        ws2 = wb.create_sheet("Computed DLT")
        _write_header_row(ws2, ["Part Number", "Manual DLT", "Computed DLT", "Δ (days)", "Critical Path"])
        for i, r in enumerate(dlt_map.values(), start=2):
            row_data = [
                r.part_number, r.manual_dlt, round(r.computed_dlt, 2),
                round(r.computed_dlt - r.manual_dlt, 2),
                " → ".join(r.critical_path),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                ws2.cell(row=i, column=col_idx, value=val).border = THIN_BORDER

        _autofit_columns(ws)
        _autofit_columns(ws2)
        buf = _wb_to_bytes(wb)
        st.download_button(
            "Download BOM Report", data=buf,
            file_name=f"DDMRP_BOM_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ---------------------------------------------------------------------------
# Model Velocity export
# ---------------------------------------------------------------------------

def _export_model_velocity():
    st.subheader("Model Velocity Export")
    st.caption("Exports the Model Velocity analysis for the DDS&OP review (slide 131).")

    window = st.number_input("Review window (days)", min_value=7, max_value=180,
                              value=30, step=7, key="exp_mv_window")

    if st.button("Generate Model Velocity Report", type="primary", key="exp_mv"):
        from views.model_velocity import compute_model_velocity
        rows = compute_model_velocity(int(window))

        wb = Workbook()
        ws = wb.active
        ws.title = "Model Velocity"
        headers = [
            "Part Number", "Description", "ADU", "Green Zone",
            "Model Order Freq (days)", "Expected Orders", "Actual Orders",
            "Velocity", "Assessment",
        ]
        _write_header_row(ws, headers)

        for i, r in enumerate(rows, start=2):
            vel = r["velocity"]
            assessment = (
                "Too Fast" if vel is not None and vel > 0.5
                else ("Too Slow" if vel is not None and vel < -0.5
                      else ("On Model" if vel is not None else "N/A"))
            )
            row_data = [
                r["part_number"], r["description"],
                r["adu"], r["green"],
                r["model_freq"] if r["model_freq"] else "",
                r["expected"] if r["expected"] else "",
                r["actual"],
                f"{vel:+.2f}" if vel is not None else "",
                assessment,
            ]
            fill = PatternFill("solid", fgColor=(
                "FADBD8" if assessment == "Too Fast"
                else ("D6EAF8" if assessment == "Too Slow"
                      else ("D5F5E3" if assessment == "On Model" else "F2F3F4"))
            ))
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col_idx, value=val)
                cell.fill = fill
                cell.border = THIN_BORDER

        _autofit_columns(ws)
        buf = _wb_to_bytes(wb)
        st.download_button(
            "Download Model Velocity Report", data=buf,
            file_name=f"DDMRP_ModelVelocity_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_header_row(ws, headers, row=1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _autofit_columns(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
