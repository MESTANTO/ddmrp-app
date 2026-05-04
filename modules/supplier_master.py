"""
Supplier Master — Streamlit page.
Manage supplier records: view, add, edit, delete, bulk import/export via Excel.
Each Item can reference a default supplier; purchase orders may override it.
"""

import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from database.db import get_session, Supplier, Item

# ── Styling constants (match the rest of the app) ────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EXAMPLE_FILL = PatternFill("solid", fgColor="D6EAF8")
EXAMPLE_FONT = Font(italic=True, color="1A5276")
INSTR_FILL  = PatternFill("solid", fgColor="FDFDE7")
THIN = Side(border_style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_COLOR = {"active": "#27AE60", "preferred": "#2980B9", "inactive": "#95A5A6"}

SUPPLIER_HEADERS = [
    # (column_name, width, example_value)
    ("Supplier Code *",           16, "SUP001"),
    ("Name *",                    28, "Acme Components SpA"),
    ("Country",                   14, "Italy"),
    ("City",                      14, "Milan"),
    ("Address",                   30, "Via Roma 1, 20100 Milan"),
    ("Website",                   24, "www.acme.it"),
    ("Phone",                     16, "+39 02 1234567"),
    ("Email",                     26, "info@acme.it"),
    ("Material Contact Name",     22, "Mario Rossi"),
    ("Material Contact Email",    26, "m.rossi@acme.it"),
    ("Material Contact Phone",    18, "+39 02 1234568"),
    ("Procurement Contact Name",  22, "Luigi Bianchi"),
    ("Procurement Contact Email", 26, "l.bianchi@acme.it"),
    ("Procurement Contact Phone", 18, "+39 02 1234569"),
    ("Manager Contact Name",      22, "Anna Verdi"),
    ("Manager Contact Email",     26, "a.verdi@acme.it"),
    ("Manager Contact Phone",     18, "+39 02 1234570"),
    ("Lead Time Days",            14, 15),
    ("Reliability %",             14, 95.0),
    ("Payment Terms",             16, "Net 30"),
    ("Currency",                  10, "EUR"),
    ("Incoterms",                 12, "DDP"),
    ("Status",                    12, "active"),
    ("Certifications",            24, "ISO 9001"),
    ("Notes",                     30, "Preferred supplier for fasteners"),
]


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def show():
    st.header("🏭 Supplier Master")
    st.caption("Manage supplier records. Assign a default supplier to each item in Material Master.")

    tab_list, tab_add, tab_edit, tab_import = st.tabs([
        "📋 Supplier List", "➕ Add Supplier", "✏️ Edit / Delete", "📥 Import / Export"
    ])

    with tab_list:
        _render_list()

    with tab_add:
        _render_add()

    with tab_edit:
        _render_edit()

    with tab_import:
        _render_import_export()


# ---------------------------------------------------------------------------
# List
# ---------------------------------------------------------------------------

def _render_list():
    session = get_session()
    try:
        suppliers = session.query(Supplier).order_by(Supplier.code).all()
        rows = []
        for s in suppliers:
            item_count = len(s.items)
            rows.append({
                "Code":             s.code,
                "Name":             s.name,
                "Country":          s.country,
                "City":             s.city,
                "Lead Time (d)":    s.lead_time_days,
                "Reliability %":    f"{s.reliability_pct:.0f}%",
                "Currency":         s.currency,
                "Payment Terms":    s.payment_terms,
                "Status":           s.status,
                "Items assigned":   item_count,
                "Material Contact": s.material_contact_name,
                "Procurement":      s.procurement_contact_name,
            })
    finally:
        session.close()

    if not rows:
        st.info("No suppliers yet. Use the **Add Supplier** tab or import from Excel.")
        return

    # KPI strip
    total = len(rows)
    active    = sum(1 for r in rows if r["Status"] == "active")
    preferred = sum(1 for r in rows if r["Status"] == "preferred")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Total suppliers", total)
    k2.metric("Active",    active)
    k3.metric("Preferred", preferred)
    k4.metric("Inactive",  total - active - preferred)

    st.divider()

    df = pd.DataFrame(rows)

    def _style_status(val):
        color = STATUS_COLOR.get(val, "#FFFFFF")
        return f"background-color:{color};color:white;font-weight:bold;border-radius:4px"

    st.dataframe(
        df.style.map(_style_status, subset=["Status"]),
        use_container_width=True,
        hide_index=True,
        height=420,
    )


# ---------------------------------------------------------------------------
# Add
# ---------------------------------------------------------------------------

def _render_add():
    st.subheader("Add New Supplier")
    session = get_session()
    try:
        existing_codes = {s.code for s in session.query(Supplier.code).all()}
    finally:
        session.close()

    with st.form("add_supplier_form", clear_on_submit=True):
        st.markdown("**General**")
        c1, c2 = st.columns(2)
        code   = c1.text_input("Supplier Code *", placeholder="SUP001")
        name   = c2.text_input("Name *",          placeholder="Acme Components SpA")

        c3, c4, c5 = st.columns(3)
        country = c3.text_input("Country", placeholder="Italy")
        city    = c4.text_input("City",    placeholder="Milan")
        status  = c5.selectbox("Status", ["active", "preferred", "inactive"])
        address = st.text_input("Address", placeholder="Via Roma 1, 20100 Milan")

        c6, c7, c8 = st.columns(3)
        website = c6.text_input("Website",  placeholder="www.acme.it")
        phone   = c7.text_input("Phone",    placeholder="+39 02 1234567")
        email   = c8.text_input("Email",    placeholder="info@acme.it")

        st.divider()
        st.markdown("**Contacts**")
        cc1, cc2, cc3 = st.columns(3)
        with cc1:
            st.markdown("*Material*")
            mat_name  = st.text_input("Name",  key="mat_name",  placeholder="Mario Rossi")
            mat_email = st.text_input("Email", key="mat_email", placeholder="m.rossi@acme.it")
            mat_phone = st.text_input("Phone", key="mat_phone", placeholder="+39 02 …")
        with cc2:
            st.markdown("*Procurement*")
            pro_name  = st.text_input("Name",  key="pro_name",  placeholder="Luigi Bianchi")
            pro_email = st.text_input("Email", key="pro_email", placeholder="l.bianchi@acme.it")
            pro_phone = st.text_input("Phone", key="pro_phone", placeholder="+39 02 …")
        with cc3:
            st.markdown("*Manager*")
            mgr_name  = st.text_input("Name",  key="mgr_name",  placeholder="Anna Verdi")
            mgr_email = st.text_input("Email", key="mgr_email", placeholder="a.verdi@acme.it")
            mgr_phone = st.text_input("Phone", key="mgr_phone", placeholder="+39 02 …")

        st.divider()
        st.markdown("**Procurement Parameters**")
        p1, p2, p3, p4, p5 = st.columns(5)
        lead_time    = p1.number_input("Lead Time (days)", min_value=0, value=0)
        reliability  = p2.number_input("Reliability %",   min_value=0.0, max_value=100.0, value=100.0, step=1.0)
        payment      = p3.text_input("Payment Terms", placeholder="Net 30")
        currency     = p4.text_input("Currency",      value="EUR")
        incoterms    = p5.text_input("Incoterms",     placeholder="DDP")

        certifications = st.text_input("Certifications", placeholder="ISO 9001, ISO 14001")
        notes          = st.text_area("Notes", height=60)

        submitted = st.form_submit_button("➕ Add Supplier", type="primary")

    if submitted:
        if not code.strip() or not name.strip():
            st.error("Supplier Code and Name are required.")
            return
        if code.strip() in existing_codes:
            st.error(f"Supplier code **{code}** already exists.")
            return
        session = get_session()
        try:
            session.add(Supplier(
                code=code.strip(), name=name.strip(),
                country=country, city=city, address=address,
                website=website, phone=phone, email=email,
                material_contact_name=mat_name,   material_contact_email=mat_email,
                material_contact_phone=mat_phone,
                procurement_contact_name=pro_name, procurement_contact_email=pro_email,
                procurement_contact_phone=pro_phone,
                manager_contact_name=mgr_name,    manager_contact_email=mgr_email,
                manager_contact_phone=mgr_phone,
                lead_time_days=int(lead_time), reliability_pct=reliability,
                payment_terms=payment, currency=currency, incoterms=incoterms,
                status=status, certifications=certifications, notes=notes,
            ))
            session.commit()
            st.success(f"Supplier **{code}** added.")
            st.rerun()
        except Exception as exc:
            session.rollback()
            st.error(f"Error: {exc}")
        finally:
            session.close()


# ---------------------------------------------------------------------------
# Edit / Delete
# ---------------------------------------------------------------------------

def _render_edit():
    session = get_session()
    try:
        suppliers = session.query(Supplier).order_by(Supplier.code).all()
        sup_opts  = {f"{s.code} — {s.name}": s.id for s in suppliers}
    finally:
        session.close()

    if not sup_opts:
        st.info("No suppliers to edit.")
        return

    sel_label = st.selectbox("Select supplier", list(sup_opts.keys()), key="edit_sup_sel")
    sel_id    = sup_opts[sel_label]

    session = get_session()
    try:
        s = session.query(Supplier).get(sel_id)
        if not s:
            st.error("Supplier not found.")
            return
        data = {col.name: getattr(s, col.name)
                for col in Supplier.__table__.columns
                if col.name not in ("id", "created_at", "updated_at")}
    finally:
        session.close()

    with st.form("edit_supplier_form"):
        st.markdown("**General**")
        c1, c2 = st.columns(2)
        code   = c1.text_input("Supplier Code *", value=data["code"])
        name   = c2.text_input("Name *",          value=data["name"])

        c3, c4, c5 = st.columns(3)
        country = c3.text_input("Country", value=data["country"] or "")
        city    = c4.text_input("City",    value=data["city"] or "")
        status  = c5.selectbox("Status", ["active", "preferred", "inactive"],
                               index=["active","preferred","inactive"].index(data["status"] or "active"))
        address = st.text_input("Address", value=data["address"] or "")

        c6, c7, c8 = st.columns(3)
        website = c6.text_input("Website", value=data["website"] or "")
        phone   = c7.text_input("Phone",   value=data["phone"]   or "")
        email   = c8.text_input("Email",   value=data["email"]   or "")

        st.divider()
        st.markdown("**Contacts**")
        cc1, cc2, cc3 = st.columns(3)
        with cc1:
            st.markdown("*Material*")
            mat_name  = st.text_input("Name",  value=data["material_contact_name"]  or "", key="emat_n")
            mat_email = st.text_input("Email", value=data["material_contact_email"] or "", key="emat_e")
            mat_phone = st.text_input("Phone", value=data["material_contact_phone"] or "", key="emat_p")
        with cc2:
            st.markdown("*Procurement*")
            pro_name  = st.text_input("Name",  value=data["procurement_contact_name"]  or "", key="epro_n")
            pro_email = st.text_input("Email", value=data["procurement_contact_email"] or "", key="epro_e")
            pro_phone = st.text_input("Phone", value=data["procurement_contact_phone"] or "", key="epro_p")
        with cc3:
            st.markdown("*Manager*")
            mgr_name  = st.text_input("Name",  value=data["manager_contact_name"]  or "", key="emgr_n")
            mgr_email = st.text_input("Email", value=data["manager_contact_email"] or "", key="emgr_e")
            mgr_phone = st.text_input("Phone", value=data["manager_contact_phone"] or "", key="emgr_p")

        st.divider()
        st.markdown("**Procurement Parameters**")
        p1, p2, p3, p4, p5 = st.columns(5)
        lead_time   = p1.number_input("Lead Time (days)", min_value=0,   value=int(data["lead_time_days"] or 0))
        reliability = p2.number_input("Reliability %",   min_value=0.0, max_value=100.0,
                                      value=float(data["reliability_pct"] or 100.0), step=1.0)
        payment     = p3.text_input("Payment Terms", value=data["payment_terms"] or "")
        currency    = p4.text_input("Currency",      value=data["currency"]      or "EUR")
        incoterms   = p5.text_input("Incoterms",     value=data["incoterms"]     or "")

        certifications = st.text_input("Certifications", value=data["certifications"] or "")
        notes          = st.text_area("Notes", value=data["notes"] or "", height=60)

        col_save, col_del = st.columns(2)
        save = col_save.form_submit_button("💾 Save Changes", type="primary")
        delete = col_del.form_submit_button("🗑️ Delete Supplier", type="secondary")

    if save:
        session = get_session()
        try:
            s = session.query(Supplier).get(sel_id)
            s.code=code.strip(); s.name=name.strip()
            s.country=country; s.city=city; s.address=address
            s.website=website; s.phone=phone; s.email=email
            s.material_contact_name=mat_name;   s.material_contact_email=mat_email;  s.material_contact_phone=mat_phone
            s.procurement_contact_name=pro_name; s.procurement_contact_email=pro_email; s.procurement_contact_phone=pro_phone
            s.manager_contact_name=mgr_name;    s.manager_contact_email=mgr_email;   s.manager_contact_phone=mgr_phone
            s.lead_time_days=int(lead_time); s.reliability_pct=reliability
            s.payment_terms=payment; s.currency=currency; s.incoterms=incoterms
            s.status=status; s.certifications=certifications; s.notes=notes
            s.updated_at=datetime.utcnow()
            session.commit()
            st.success("Saved.")
            st.rerun()
        except Exception as exc:
            session.rollback(); st.error(str(exc))
        finally:
            session.close()

    if delete:
        session = get_session()
        try:
            # unlink items first
            for item in session.query(Item).filter_by(default_supplier_id=sel_id).all():
                item.default_supplier_id = None
            session.query(Supplier).filter_by(id=sel_id).delete()
            session.commit()
            st.success("Supplier deleted.")
            st.rerun()
        except Exception as exc:
            session.rollback(); st.error(str(exc))
        finally:
            session.close()


# ---------------------------------------------------------------------------
# Import / Export
# ---------------------------------------------------------------------------

def _render_import_export():
    col_imp, col_exp = st.columns(2)

    with col_imp:
        st.subheader("📥 Import from Excel")
        tmpl_bytes = _build_template()
        st.download_button(
            "⬇️ Download template",
            data=tmpl_bytes,
            file_name="DDMRP_Suppliers_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        uploaded = st.file_uploader("Upload filled template", type=["xlsx"], key="sup_upload")
        if uploaded:
            ok, errors = _import_suppliers(uploaded)
            if ok:
                st.success(f"✅ {ok} supplier(s) imported.")
            if errors:
                st.error("\n".join(errors[:20]))

    with col_exp:
        st.subheader("📤 Export to Excel")
        if st.button("⬇️ Export all suppliers", type="primary"):
            data = _export_suppliers()
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
            st.download_button(
                "Download",
                data=data,
                file_name=f"DDMRP_Suppliers_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


# ---------------------------------------------------------------------------
# Excel template builder
# ---------------------------------------------------------------------------

def _build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    # Header row
    for col_idx, (col_name, width, _) in enumerate(SUPPLIER_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Example row
    for col_idx, (_, _, example) in enumerate(SUPPLIER_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=example)
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER

    # Instructions row
    instr = ws.cell(row=3, column=1,
                    value="← Example row above. Add your data from row 4 onwards. "
                          "Status: active | preferred | inactive")
    instr.fill = INSTR_FILL
    instr.font = Font(italic=True, color="856404")
    ws.merge_cells(start_row=3, start_column=1,
                   end_row=3, end_column=len(SUPPLIER_HEADERS))
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel import
# ---------------------------------------------------------------------------

def _import_suppliers(file) -> tuple[int, list[str]]:
    try:
        df = pd.read_excel(file, header=0, skiprows=[1, 2])
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")
    except Exception as e:
        return 0, [f"Cannot read file: {e}"]

    col_map = {h: h.replace(" *", "") for h, _, _ in SUPPLIER_HEADERS}

    def _get(row, col, default=""):
        for key in (col, col + " *"):
            if key in row.index:
                v = row[key]
                return "" if pd.isna(v) else str(v).strip()
        return default

    session = get_session()
    ok = 0
    errors = []
    try:
        existing = {s.code: s for s in session.query(Supplier).all()}
        for i, row in df.iterrows():
            code = _get(row, "Supplier Code")
            name = _get(row, "Name")
            if not code or not name:
                errors.append(f"Row {i+4}: missing Supplier Code or Name — skipped.")
                continue

            s = existing.get(code) or Supplier(code=code)
            s.name                    = name
            s.country                 = _get(row, "Country")
            s.city                    = _get(row, "City")
            s.address                 = _get(row, "Address")
            s.website                 = _get(row, "Website")
            s.phone                   = _get(row, "Phone")
            s.email                   = _get(row, "Email")
            s.material_contact_name   = _get(row, "Material Contact Name")
            s.material_contact_email  = _get(row, "Material Contact Email")
            s.material_contact_phone  = _get(row, "Material Contact Phone")
            s.procurement_contact_name  = _get(row, "Procurement Contact Name")
            s.procurement_contact_email = _get(row, "Procurement Contact Email")
            s.procurement_contact_phone = _get(row, "Procurement Contact Phone")
            s.manager_contact_name    = _get(row, "Manager Contact Name")
            s.manager_contact_email   = _get(row, "Manager Contact Email")
            s.manager_contact_phone   = _get(row, "Manager Contact Phone")
            try:
                s.lead_time_days  = int(float(_get(row, "Lead Time Days") or 0))
                s.reliability_pct = float(_get(row, "Reliability %") or 100.0)
            except ValueError:
                s.lead_time_days = 0; s.reliability_pct = 100.0
            s.payment_terms   = _get(row, "Payment Terms")
            s.currency        = _get(row, "Currency") or "EUR"
            s.incoterms       = _get(row, "Incoterms")
            raw_status        = _get(row, "Status").lower()
            s.status          = raw_status if raw_status in ("active","preferred","inactive") else "active"
            s.certifications  = _get(row, "Certifications")
            s.notes           = _get(row, "Notes")
            s.updated_at      = datetime.utcnow()

            if code not in existing:
                session.add(s)
                existing[code] = s
            ok += 1

        session.commit()
    except Exception as e:
        session.rollback()
        errors.append(f"DB error: {e}")
    finally:
        session.close()

    return ok, errors


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _export_suppliers() -> bytes:
    session = get_session()
    try:
        suppliers = session.query(Supplier).order_by(Supplier.code).all()
    finally:
        session.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    headers = [h.replace(" *", "") for h, _, _ in SUPPLIER_HEADERS] + ["Created At", "Updated At"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            SUPPLIER_HEADERS[col_idx-1][1] if col_idx <= len(SUPPLIER_HEADERS) else 18
        )

    for row_idx, s in enumerate(suppliers, 2):
        vals = [
            s.code, s.name, s.country, s.city, s.address,
            s.website, s.phone, s.email,
            s.material_contact_name, s.material_contact_email, s.material_contact_phone,
            s.procurement_contact_name, s.procurement_contact_email, s.procurement_contact_phone,
            s.manager_contact_name, s.manager_contact_email, s.manager_contact_phone,
            s.lead_time_days, s.reliability_pct, s.payment_terms,
            s.currency, s.incoterms, s.status, s.certifications, s.notes,
            s.created_at, s.updated_at,
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
