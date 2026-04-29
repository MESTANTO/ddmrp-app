"""
Shared Excel import utilities.
Each import function returns (success_count, error_list).
Template generators return bytes ready for st.download_button.
"""

import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL  = PatternFill("solid", fgColor="2C3E50")
EXAMPLE_FILL = PatternFill("solid", fgColor="EBF5FB")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
EXAMPLE_FONT = Font(color="1A5276", italic=True)
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_header(ws, headers: list[dict], row: int = 1):
    """
    headers: list of dicts with keys: name, width (optional), note (optional)
    """
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h["name"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = h.get("width", 18)
    ws.row_dimensions[row].height = 28


def _write_example_row(ws, values: list, row: int = 2):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _add_instructions(ws, text: str, row: int, col_span: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = PatternFill("solid", fgColor="FEF9E7")
    cell.font = Font(italic=True, color="7D6608", size=9)
    cell.alignment = Alignment(wrap_text=True)


def _read_uploaded_file(uploaded_file):
    """Read an uploaded Streamlit file into a DataFrame. Returns None on error."""
    try:
        df = pd.read_excel(uploaded_file, header=0, skiprows=2)  # skip header + example row
        df.columns = df.columns.str.strip()
        df = df.dropna(how="all")
        return df
    except Exception as e:
        return None


# ---------------------------------------------------------------------------
# Material Master template + importer
# ---------------------------------------------------------------------------

MATERIAL_HEADERS = [
    {"name": "Part Number *",        "width": 16},
    {"name": "Description *",        "width": 28},
    {"name": "Category",             "width": 16},
    {"name": "Unit of Measure",      "width": 14},
    {"name": "Item Type",            "width": 12},
    {"name": "Buffer Profile",       "width": 16},
    {"name": "ADU *",                "width": 12},
    {"name": "DLT (days) *",         "width": 14},
    {"name": "Lead Time Factor",     "width": 16},
    {"name": "Variability Factor",   "width": 16},
    {"name": "Min Order Qty",        "width": 14},
    {"name": "Order Cycle (days)",   "width": 16},
    {"name": "Spike Horizon (days)", "width": 18},
    {"name": "Spike Threshold Factor","width": 20},
    {"name": "On Hand",              "width": 12},
    {"name": "Unit Cost (€)",        "width": 14},
    {"name": "Ordering Cost (€)",    "width": 16},
    {"name": "Holding Cost %",       "width": 16},
]

MATERIAL_EXAMPLE = [
    "RM-001", "Raw Material A", "Raw Material", "KG",
    "P", "P-M-M",
    10.0, 5.0, 0.5, 0.5, 50.0, 7.0,
    "", "",
    100.0,
    25.0, 50.0, 25.0,
]


def build_material_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    _add_instructions(ws,
        "Instructions: Fill from row 4 onward. Row 3 is an example (do not delete rows 1-3). "
        "Fields marked * are required. "
        "Item Type: M=Manufactured, I=Intermediate, P=Purchased, D=Distributed (slide 50). "
        "Buffer Profile is the Item Type-LeadTimeCategory-VariabilityCategory code, e.g. 'P-M-M' (slides 51-54); "
        "leave blank to use manual LTF/VF only. Bands: LTF L=0.20-0.40, M=0.41-0.60, S=0.61-1.0; "
        "VF L=0.0-0.40, M=0.41-0.60, H=0.61-1.0. "
        "Spike Horizon (days) and Spike Threshold Factor (× ADU) are ASOH overrides (slide 83) — "
        "leave blank to use global defaults. "
        "Unit of Measure options: EA, KG, LT, M, MT, PC. "
        "Cost fields are optional. Holding Cost % is annual percentage (e.g. 25 = 25%).",
        row=1, col_span=len(MATERIAL_HEADERS))
    _write_header(ws, MATERIAL_HEADERS, row=2)
    _write_example_row(ws, MATERIAL_EXAMPLE, row=3)

    # Resolve column letters by header name to keep dropdowns in sync if columns shift
    name_to_col = {h["name"]: get_column_letter(i + 1)
                   for i, h in enumerate(MATERIAL_HEADERS)}

    # Dropdown validation for UoM
    dv_uom = DataValidation(type="list", formula1='"EA,KG,LT,M,MT,PC"', allow_blank=True)
    ws.add_data_validation(dv_uom)
    dv_uom.sqref = f'{name_to_col["Unit of Measure"]}4:{name_to_col["Unit of Measure"]}1000'

    # Dropdown validation for Item Type
    dv_type = DataValidation(type="list", formula1='"M,I,P,D"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.sqref = f'{name_to_col["Item Type"]}4:{name_to_col["Item Type"]}1000'

    # Dropdown validation for LTF
    dv_ltf = DataValidation(type="list",
                            formula1='"0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9,1.0"',
                            allow_blank=True)
    ws.add_data_validation(dv_ltf)
    dv_ltf.sqref = f'{name_to_col["Lead Time Factor"]}4:{name_to_col["Lead Time Factor"]}1000'

    # Dropdown validation for VF
    dv_vf = DataValidation(type="list",
                           formula1='"0.0,0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9,1.0"',
                           allow_blank=True)
    ws.add_data_validation(dv_vf)
    dv_vf.sqref = f'{name_to_col["Variability Factor"]}4:{name_to_col["Variability Factor"]}1000'

    ws.freeze_panes = "A4"
    return _wb_to_bytes(wb)


def import_materials(uploaded_file) -> tuple[int, list[str]]:
    from database.db import get_session, Item, BufferProfile

    # Canonical bands for LTF / VF validation (deck slides 51-54)
    LTF_BANDS = {"S": (0.61, 1.00), "M": (0.41, 0.60), "L": (0.20, 0.40)}
    VF_BANDS  = {"L": (0.00, 0.40), "M": (0.41, 0.60), "H": (0.61, 1.00)}
    VALID_TYPES = {"M", "I", "P", "D"}

    try:
        df = pd.read_excel(uploaded_file, header=1, skiprows=[2])  # header at row 2, skip row 3 (example)
    except Exception as e:
        return 0, [f"Could not read file: {e}"]

    df.columns = df.columns.str.strip()
    df = df.dropna(how="all")
    df.columns = [c.replace(" *", "").strip() for c in df.columns]

    # Pre-load profiles
    session = get_session()
    try:
        profile_map = {p.name: p for p in session.query(BufferProfile).all()}
    finally:
        session.close()

    errors = []
    success = 0

    def _opt_int(v):
        if v is None or pd.isna(v) or str(v).strip() == "":
            return None
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None

    def _opt_float(v):
        if v is None or pd.isna(v) or str(v).strip() == "":
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    # --- Pre-validate all rows before touching the DB ---
    new_items = []
    for idx, row in df.iterrows():
        row_num = idx + 4
        part = str(row.get("Part Number", "")).strip().upper()
        desc = str(row.get("Description", "")).strip()
        if not part or part == "NAN":
            continue
        if not desc or desc == "nan":
            errors.append(f"Row {row_num}: Description is required for {part}.")
            continue

        try:
            ltf = float(row.get("Lead Time Factor", 0.5) or 0.5)
            vf  = float(row.get("Variability Factor", 0.5) or 0.5)

            item_type = str(row.get("Item Type", "P") or "P").strip().upper()
            if item_type not in VALID_TYPES:
                item_type = "P"

            raw_profile = row.get("Buffer Profile", "")
            profile_name = ("" if pd.isna(raw_profile) else str(raw_profile)).strip()
            if profile_name.lower() == "nan":
                profile_name = ""
            profile_obj  = profile_map.get(profile_name) if profile_name else None
            if profile_name and profile_obj is None:
                errors.append(f"Row {row_num} ({part}): Buffer Profile '{profile_name}' not found "
                              "(expected format e.g. 'P-M-M'). Profile cleared.")
                profile_name = ""

            # Band validation if a profile is set
            if profile_obj is not None:
                ltf_band = LTF_BANDS.get(profile_obj.lt_category)
                vf_band  = VF_BANDS.get(profile_obj.var_category)
                if ltf_band and not (ltf_band[0] <= ltf <= ltf_band[1]):
                    errors.append(f"Row {row_num} ({part}): LTF {ltf} outside band "
                                  f"{ltf_band[0]:.2f}-{ltf_band[1]:.2f} for category "
                                  f"{profile_obj.lt_category}. Row skipped.")
                    continue
                if vf_band and not (vf_band[0] <= vf <= vf_band[1]):
                    errors.append(f"Row {row_num} ({part}): VF {vf} outside band "
                                  f"{vf_band[0]:.2f}-{vf_band[1]:.2f} for category "
                                  f"{profile_obj.var_category}. Row skipped.")
                    continue

            new_items.append(dict(
                part_number = part,
                description = desc,
                category    = str(row.get("Category", "") or "").strip(),
                unit_of_measure = str(row.get("Unit of Measure", "EA") or "EA").strip(),
                item_type   = item_type,
                buffer_profile_id = profile_obj.id if profile_obj else None,
                adu  = float(row.get("ADU", 0) or 0),
                dlt  = float(row.get("DLT (days)", 0) or 0),
                lead_time_factor   = ltf,
                variability_factor = vf,
                min_order_qty = float(row.get("Min Order Qty", 0) or 0),
                order_cycle   = float(row.get("Order Cycle (days)", 0) or 0),
                spike_horizon_days     = _opt_int(row.get("Spike Horizon (days)")),
                spike_threshold_factor = _opt_float(row.get("Spike Threshold Factor")),
                on_hand       = float(row.get("On Hand", 0) or 0),
                unit_cost        = float(row.get("Unit Cost (€)", 0) or 0),
                ordering_cost    = float(row.get("Ordering Cost (€)", 0) or 0),
                holding_cost_pct = float(row.get("Holding Cost %", 0) or 0) / 100.0,
            ))
        except (ValueError, TypeError) as e:
            errors.append(f"Row {row_num} ({part}): Invalid numeric value — {e}")

    if not new_items:
        return 0, errors + ["No valid rows found in file — existing data was NOT deleted."]

    session = get_session()
    try:
        # DELETE all existing items (cascades to demand, supply, buffers)
        session.query(Item).delete()
        session.flush()

        # INSERT fresh data from file
        for d in new_items:
            uom = d["unit_of_measure"] if d["unit_of_measure"] in ["EA","KG","LT","M","MT","PC"] else "EA"
            session.add(Item(
                part_number=d["part_number"], description=d["description"],
                category=d["category"], unit_of_measure=uom,
                item_type=d["item_type"],
                buffer_profile_id=d["buffer_profile_id"],
                adu=d["adu"], dlt=d["dlt"],
                lead_time_factor=d["lead_time_factor"],
                variability_factor=d["variability_factor"],
                min_order_qty=d["min_order_qty"],
                order_cycle=d["order_cycle"],
                spike_horizon_days=d["spike_horizon_days"],
                spike_threshold_factor=d["spike_threshold_factor"],
                on_hand=d["on_hand"],
                unit_cost=d["unit_cost"],
                ordering_cost=d["ordering_cost"],
                holding_cost_pct=d["holding_cost_pct"],
            ))
            success += 1

        session.commit()
    except Exception as e:
        session.rollback()
        errors.append(f"Database error: {e}")
        success = 0
    finally:
        session.close()

    return success, errors


# ---------------------------------------------------------------------------
# Demand template + importer
# ---------------------------------------------------------------------------

DEMAND_HEADERS = [
    {"name": "Part Number *",   "width": 16},
    {"name": "Demand Type *",   "width": 14},
    {"name": "Quantity *",      "width": 12},
    {"name": "Demand Date *",   "width": 16},
    {"name": "Order Reference", "width": 18},
    {"name": "Notes",           "width": 28},
]

DEMAND_EXAMPLE = ["RM-001", "actual", 50.0, "2026-04-01", "SO-12345", "Optional notes"]


def build_demand_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Demand"

    _add_instructions(ws,
        "Instructions: Fill from row 4 onward. Demand Type: 'actual' or 'forecast'. "
        "Date format: YYYY-MM-DD. Part Number must exist in Material Master.",
        row=1, col_span=len(DEMAND_HEADERS))
    _write_header(ws, DEMAND_HEADERS, row=2)
    _write_example_row(ws, DEMAND_EXAMPLE, row=3)

    dv = DataValidation(type="list", formula1='"actual,forecast"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.sqref = "B4:B1000"

    ws.freeze_panes = "A4"
    return _wb_to_bytes(wb)


def import_demand(uploaded_file) -> tuple[int, list[str]]:
    from database.db import get_session, Item, DemandEntry

    try:
        df = pd.read_excel(uploaded_file, header=1, skiprows=[2])
    except Exception as e:
        return 0, [f"Could not read file: {e}"]

    df.columns = [c.replace(" *", "").strip() for c in df.columns]
    df = df.dropna(how="all")

    errors = []
    new_entries = []

    # --- Pre-validate ---
    session = get_session()
    try:
        item_map = {it.part_number: it.id for it in session.query(Item).all()}
    finally:
        session.close()

    for idx, row in df.iterrows():
        row_num = idx + 4
        part = str(row.get("Part Number", "")).strip().upper()
        if not part or part == "NAN":
            continue
        if part not in item_map:
            errors.append(f"Row {row_num}: Part '{part}' not found in Material Master.")
            continue
        try:
            qty   = float(row.get("Quantity", 0))
            dtype = str(row.get("Demand Type", "actual")).strip().lower()
            if dtype not in ("actual", "forecast"):
                dtype = "actual"
            raw_date = row.get("Demand Date")
            if pd.isna(raw_date):
                errors.append(f"Row {row_num} ({part}): Demand Date is required.")
                continue
            new_entries.append(dict(
                item_id=item_map[part],
                demand_type=dtype,
                quantity=qty,
                demand_date=pd.to_datetime(raw_date).to_pydatetime(),
                order_reference=str(row.get("Order Reference", "") or "").strip(),
                notes=str(row.get("Notes", "") or "").strip(),
            ))
        except Exception as e:
            errors.append(f"Row {row_num} ({part}): {e}")

    if not new_entries:
        return 0, errors + ["No valid rows found — existing data was NOT deleted."]

    session = get_session()
    try:
        # DELETE all existing demand entries
        session.query(DemandEntry).delete()
        session.flush()

        # INSERT fresh rows
        for d in new_entries:
            session.add(DemandEntry(**d))

        session.commit()
        return len(new_entries), errors
    except Exception as e:
        session.rollback()
        return 0, errors + [f"Database error: {e}"]
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Supply template + importer
# ---------------------------------------------------------------------------

SUPPLY_HEADERS = [
    {"name": "Part Number *",   "width": 16},
    {"name": "Supply Type *",   "width": 20},
    {"name": "Quantity *",      "width": 12},
    {"name": "Due Date *",      "width": 16},
    {"name": "Order Reference", "width": 18},
    {"name": "Notes",           "width": 28},
]

SUPPLY_EXAMPLE = ["RM-001", "purchase_order", 200.0, "2026-04-10", "PO-9876", ""]


def build_supply_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Supply"

    _add_instructions(ws,
        "Instructions: Fill from row 4 onward. Supply Type: 'purchase_order' or 'production_order'. "
        "Date format: YYYY-MM-DD. Part Number must exist in Material Master.",
        row=1, col_span=len(SUPPLY_HEADERS))
    _write_header(ws, SUPPLY_HEADERS, row=2)
    _write_example_row(ws, SUPPLY_EXAMPLE, row=3)

    dv = DataValidation(type="list",
                        formula1='"purchase_order,production_order"',
                        allow_blank=False)
    ws.add_data_validation(dv)
    dv.sqref = "B4:B1000"

    ws.freeze_panes = "A4"
    return _wb_to_bytes(wb)


def import_supply(uploaded_file) -> tuple[int, list[str]]:
    from database.db import get_session, Item, SupplyEntry

    try:
        df = pd.read_excel(uploaded_file, header=1, skiprows=[2])
    except Exception as e:
        return 0, [f"Could not read file: {e}"]

    df.columns = [c.replace(" *", "").strip() for c in df.columns]
    df = df.dropna(how="all")

    errors = []
    new_entries = []

    # --- Pre-validate ---
    session = get_session()
    try:
        item_map = {it.part_number: it.id for it in session.query(Item).all()}
    finally:
        session.close()

    for idx, row in df.iterrows():
        row_num = idx + 4
        part = str(row.get("Part Number", "")).strip().upper()
        if not part or part == "NAN":
            continue
        if part not in item_map:
            errors.append(f"Row {row_num}: Part '{part}' not found in Material Master.")
            continue
        try:
            qty   = float(row.get("Quantity", 0))
            stype = str(row.get("Supply Type", "purchase_order")).strip().lower()
            if stype not in ("purchase_order", "production_order"):
                stype = "purchase_order"
            raw_date = row.get("Due Date")
            if pd.isna(raw_date):
                errors.append(f"Row {row_num} ({part}): Due Date is required.")
                continue
            new_entries.append(dict(
                item_id=item_map[part],
                supply_type=stype,
                quantity=qty,
                due_date=pd.to_datetime(raw_date).to_pydatetime(),
                order_reference=str(row.get("Order Reference", "") or "").strip(),
                notes=str(row.get("Notes", "") or "").strip(),
            ))
        except Exception as e:
            errors.append(f"Row {row_num} ({part}): {e}")

    if not new_entries:
        return 0, errors + ["No valid rows found — existing data was NOT deleted."]

    session = get_session()
    try:
        # DELETE all existing supply entries
        session.query(SupplyEntry).delete()
        session.flush()

        # INSERT fresh rows
        for d in new_entries:
            session.add(SupplyEntry(**d))

        session.commit()
        return len(new_entries), errors
    except Exception as e:
        session.rollback()
        return 0, errors + [f"Database error: {e}"]
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Process Nodes template + importer
# ---------------------------------------------------------------------------

PROCESS_NODE_HEADERS = [
    {"name": "Process Name *",      "width": 20},
    {"name": "Sequence *",          "width": 12},
    {"name": "Node Label *",        "width": 24},
    {"name": "Node Type *",         "width": 14},
    {"name": "Has Buffer (YES/NO)", "width": 18},
    {"name": "Linked Part Number",  "width": 20},
]

PROCESS_NODE_EXAMPLE = ["Assembly Line A", 1, "Cutting", "operation", "NO", "RM-001"]


def build_process_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Process Nodes"

    _add_instructions(ws,
        "Instructions: Fill from row 4 onward. Process Name must match an existing process "
        "(or a new one will be created). Node Type: operation / material / buffer. "
        "Has Buffer: YES or NO. Linked Part Number is optional.",
        row=1, col_span=len(PROCESS_NODE_HEADERS))
    _write_header(ws, PROCESS_NODE_HEADERS, row=2)
    _write_example_row(ws, PROCESS_NODE_EXAMPLE, row=3)

    dv_type = DataValidation(type="list", formula1='"operation,material,buffer"', allow_blank=False)
    ws.add_data_validation(dv_type)
    dv_type.sqref = "D4:D1000"

    dv_buf = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    ws.add_data_validation(dv_buf)
    dv_buf.sqref = "E4:E1000"

    ws.freeze_panes = "A4"
    return _wb_to_bytes(wb)


def import_process_nodes(uploaded_file) -> tuple[int, list[str]]:
    from database.db import get_session, Item, Process, ProcessNode, ProcessEdge

    try:
        df = pd.read_excel(uploaded_file, header=1, skiprows=[2])
    except Exception as e:
        return 0, [f"Could not read file: {e}"]

    df.columns = [c.replace(" *", "").strip() for c in df.columns]
    df = df.dropna(how="all")

    errors = []
    new_nodes = []

    # --- Pre-validate ---
    session = get_session()
    try:
        item_map = {it.part_number: it.id for it in session.query(Item).all()}
    finally:
        session.close()

    # Collect process names referenced in the file
    referenced_processes = set()
    for idx, row in df.iterrows():
        row_num = idx + 4
        proc_name = str(row.get("Process Name", "")).strip()
        label     = str(row.get("Node Label", "")).strip()
        if not proc_name or proc_name == "nan":
            continue
        if not label or label == "nan":
            errors.append(f"Row {row_num}: Node Label is required.")
            continue
        try:
            seq     = int(row.get("Sequence", 0) or 0)
            ntype   = str(row.get("Node Type", "operation") or "operation").strip().lower()
            if ntype not in ("operation", "material", "buffer"):
                ntype = "operation"
            has_buf = str(row.get("Has Buffer (YES/NO)", "NO")).strip().upper() == "YES"
            part_raw= str(row.get("Linked Part Number", "") or "").strip().upper()
        except Exception as e:
            errors.append(f"Row {row_num}: {e}")
            continue

        item_id = None
        if part_raw and part_raw != "NAN":
            if part_raw in item_map:
                item_id = item_map[part_raw]
            else:
                errors.append(f"Row {row_num}: Part '{part_raw}' not found — node will have no item link.")

        referenced_processes.add(proc_name)
        new_nodes.append(dict(
            proc_name=proc_name, label=label, node_type=ntype,
            has_buffer=has_buf, sequence=seq, item_id=item_id,
        ))

    if not new_nodes:
        return 0, errors + ["No valid rows found — existing data was NOT deleted."]

    session = get_session()
    try:
        # DELETE all nodes and edges for processes referenced in the file
        # (edges cascade-delete via ProcessNode relationship)
        for proc_name in referenced_processes:
            proc = session.query(Process).filter_by(name=proc_name).first()
            if proc:
                # Delete edges first to avoid FK issues
                session.query(ProcessEdge).filter_by(process_id=proc.id).delete()
                session.query(ProcessNode).filter_by(process_id=proc.id).delete()
                session.flush()

        # Re-fetch or create processes after deletion
        for node in new_nodes:
            proc = session.query(Process).filter_by(name=node["proc_name"]).first()
            if not proc:
                proc = Process(name=node["proc_name"])
                session.add(proc)
                session.flush()
            node["process_id"] = proc.id

        # INSERT fresh nodes
        for node in new_nodes:
            session.add(ProcessNode(
                process_id=node["process_id"],
                item_id=node["item_id"],
                label=node["label"],
                node_type=node["node_type"],
                has_buffer=node["has_buffer"],
                sequence=node["sequence"],
            ))

        session.commit()
        return len(new_nodes), errors
    except Exception as e:
        session.rollback()
        return 0, errors + [f"Database error: {e}"]
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Shared Streamlit UI widget (reusable across pages)
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# BOM Lines template + importer
# ---------------------------------------------------------------------------

BOM_HEADERS = [
    {"name": "Parent Part Number *", "width": 20},
    {"name": "Child Part Number *",  "width": 20},
    {"name": "Qty per Assembly *",   "width": 18},
    {"name": "Note",                 "width": 28},
]

BOM_EXAMPLE = ["FG-001", "RM-001", 2.0, "2 units of RM-001 per assembly of FG-001"]


def build_bom_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Lines"
    _add_instructions(ws,
        "Instructions: Fill from row 4 onward. Row 3 is an example. "
        "Parent Part Number = the assembly (output item). "
        "Child Part Number = the component (input item). "
        "Both part numbers must already exist in Material Master. "
        "Importing replaces ALL existing BOM lines.",
        row=1, col_span=len(BOM_HEADERS))
    _write_header(ws, BOM_HEADERS, row=2)
    _write_example_row(ws, BOM_EXAMPLE, row=3)
    ws.freeze_panes = "A4"
    return _wb_to_bytes(wb)


def import_bom(uploaded_file) -> tuple[int, list[str]]:
    """
    Replace all BomLine rows with the contents of the uploaded template.
    Returns (success_count, error_list).
    """
    from database.db import get_session, Item, BomLine

    df = _read_uploaded_file(uploaded_file)
    if df is None:
        return 0, ["Could not read file — check it uses the official template."]

    session = get_session()
    errors: list[str] = []
    count = 0
    try:
        # Wipe existing BOM lines
        session.query(BomLine).delete()
        session.commit()

        items_by_pn = {it.part_number.strip(): it for it in session.query(Item).all()}

        for i, row in df.iterrows():
            row_num = i + 4
            parent_pn = str(row.get("Parent Part Number *", "") or "").strip()
            child_pn  = str(row.get("Child Part Number *",  "") or "").strip()
            qty_raw   = row.get("Qty per Assembly *", 1.0)

            if not parent_pn or not child_pn:
                errors.append(f"Row {row_num}: missing parent or child part number — skipped.")
                continue

            parent = items_by_pn.get(parent_pn)
            child  = items_by_pn.get(child_pn)
            if not parent:
                errors.append(f"Row {row_num}: parent '{parent_pn}' not in Material Master — skipped.")
                continue
            if not child:
                errors.append(f"Row {row_num}: child '{child_pn}' not in Material Master — skipped.")
                continue
            if parent.id == child.id:
                errors.append(f"Row {row_num}: parent and child are the same item — skipped.")
                continue

            try:
                qty = float(qty_raw)
            except (TypeError, ValueError):
                qty = 1.0

            note = str(row.get("Note", "") or "").strip()
            session.add(BomLine(parent_item_id=parent.id, child_item_id=child.id,
                                qty=qty, note=note))
            count += 1

        session.commit()
    except Exception as e:
        session.rollback()
        return 0, [f"Import failed: {e}"]
    finally:
        session.close()

    return count, errors


# ---------------------------------------------------------------------------
# Buffer Adjustments template + importer
# ---------------------------------------------------------------------------

ADJ_HEADERS = [
    {"name": "Part Number *",       "width": 18},
    {"name": "Start Date * (YYYY-MM-DD)", "width": 22},
    {"name": "End Date (YYYY-MM-DD)",     "width": 22},
    {"name": "DAF",                 "width": 10},
    {"name": "LTAF",                "width": 10},
    {"name": "Red ZAF",             "width": 10},
    {"name": "Yellow ZAF",          "width": 12},
    {"name": "Green ZAF",           "width": 12},
    {"name": "Note",                "width": 28},
]

ADJ_EXAMPLE = [
    "FG-001", "2026-05-01", "2026-06-30",
    1.3, 1.0, 1.0, 1.0, 1.0,
    "Summer demand peak — DAF +30%",
]


def build_adjustments_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Buffer Adjustments"
    _add_instructions(ws,
        "Instructions: Fill from row 4 onward. Row 3 is an example. "
        "Part Number must exist in Material Master. "
        "Start Date is required; End Date blank = open-ended. "
        "DAF = Demand Adjustment Factor (× ADU). LTAF = Lead Time Adjustment Factor (× DLT). "
        "Red/Yellow/Green ZAF = Zone Adjustment Factors. "
        "All factors default to 1.0 (neutral). Importing replaces ALL existing adjustments.",
        row=1, col_span=len(ADJ_HEADERS))
    _write_header(ws, ADJ_HEADERS, row=2)
    _write_example_row(ws, ADJ_EXAMPLE, row=3)
    ws.freeze_panes = "A4"
    return _wb_to_bytes(wb)


def import_adjustments(uploaded_file) -> tuple[int, list[str]]:
    """
    Replace all BufferAdjustment rows with the contents of the uploaded template.
    Returns (success_count, error_list).
    """
    from database.db import get_session, Item, BufferAdjustment

    df = _read_uploaded_file(uploaded_file)
    if df is None:
        return 0, ["Could not read file — check it uses the official template."]

    session = get_session()
    errors: list[str] = []
    count = 0
    try:
        session.query(BufferAdjustment).delete()
        session.commit()

        items_by_pn = {it.part_number.strip(): it for it in session.query(Item).all()}

        for i, row in df.iterrows():
            row_num = i + 4
            pn = str(row.get("Part Number *", "") or "").strip()
            if not pn:
                errors.append(f"Row {row_num}: missing Part Number — skipped.")
                continue
            item = items_by_pn.get(pn)
            if not item:
                errors.append(f"Row {row_num}: '{pn}' not in Material Master — skipped.")
                continue

            start_raw = row.get("Start Date * (YYYY-MM-DD)", "")
            if not start_raw:
                errors.append(f"Row {row_num}: missing Start Date — skipped.")
                continue
            try:
                start_dt = datetime.strptime(str(start_raw).strip()[:10], "%Y-%m-%d")
            except ValueError:
                errors.append(f"Row {row_num}: invalid Start Date '{start_raw}' — skipped.")
                continue

            end_raw = row.get("End Date (YYYY-MM-DD)", "")
            end_dt  = None
            if end_raw and str(end_raw).strip():
                try:
                    end_dt = datetime.strptime(str(end_raw).strip()[:10], "%Y-%m-%d")
                except ValueError:
                    errors.append(f"Row {row_num}: invalid End Date '{end_raw}' — using open-ended.")

            def _f(col, default=1.0):
                try:
                    v = float(row.get(col, default) or default)
                    return v if v > 0 else default
                except (TypeError, ValueError):
                    return default

            session.add(BufferAdjustment(
                item_id=item.id,
                start_date=start_dt,
                end_date=end_dt,
                daf=_f("DAF"), ltaf=_f("LTAF"),
                red_zaf=_f("Red ZAF"), yellow_zaf=_f("Yellow ZAF"), green_zaf=_f("Green ZAF"),
                note=str(row.get("Note", "") or "").strip(),
            ))
            count += 1

        session.commit()
    except Exception as e:
        session.rollback()
        return 0, [f"Import failed: {e}"]
    finally:
        session.close()

    return count, errors


def render_import_widget(
    label: str,
    template_fn,
    import_fn,
    template_filename: str,
    key: str,
):
    """
    Renders a collapsible import section with:
      1. Download template button
      2. Warning that existing data will be fully replaced
      3. Confirmation checkbox
      4. File uploader
      5. Import button with result feedback
    """
    import streamlit as st

    with st.expander(f"⬆️ Import {label} from Excel", expanded=False):

        # Step 1 — Download template
        st.markdown("**Step 1 — Download the template**")
        st.download_button(
            label=f"⬇️ Download {label} Template",
            data=template_fn(),
            file_name=template_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_{key}",
        )
        st.caption("Fill in your data from row 4 onward (row 3 is a pre-filled example — do not delete it).")

        st.divider()

        # Step 2 — Warning + confirmation
        st.markdown("**Step 2 — Confirm replacement**")
        st.warning(
            f"⚠️ **Importing will permanently DELETE all existing {label} data "
            f"and replace it with the content of your file.** "
            f"This action cannot be undone.",
            icon="⚠️",
        )
        confirmed = st.checkbox(
            f"I understand — replace all existing {label} data with the imported file",
            key=f"confirm_{key}",
            value=False,
        )

        st.divider()

        # Step 3 — Upload and import
        st.markdown("**Step 3 — Upload your filled template**")
        uploaded = st.file_uploader(
            f"Upload filled {label} template (.xlsx)",
            type=["xlsx"],
            key=f"upload_{key}",
        )

        if uploaded:
            if not confirmed:
                st.info("Please tick the confirmation checkbox above to enable the import button.")
            else:
                if st.button(
                    f"🔄 Replace all {label} data with this file",
                    type="primary",
                    key=f"import_btn_{key}",
                    use_container_width=True,
                ):
                    with st.spinner(f"Deleting existing {label} data and importing new data..."):
                        count, errs = import_fn(uploaded)

                    if count:
                        st.success(f"✅ {count} row(s) imported successfully. All previous {label} data has been replaced.")
                    if errs:
                        st.warning(f"{len(errs)} row(s) had issues (these rows were skipped):")
                        for e in errs:
                            st.caption(f"  • {e}")
                    if count:
                        # Reset confirmation so user must re-tick for next import
                        st.rerun()
